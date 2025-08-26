import pandas as pd
from openpyxl import Workbook, load_workbook
from tkinter import filedialog, messagebox
import os
import tkinter as tk
import datetime
from openpyxl.styles import PatternFill
from tkinter import messagebox as mb 

def process_file(excel_file1):



    res = mb.askquestion('Exit Application',  
                         'Do you really want to exit') 
    
     
    if res == 'yes' : 
        root.destroy() 
          
    else : 
        mb.showinfo('Return', 'Returning to main application') 
        
    time = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    directory_path = os.path.dirname(excel_file1)
    sheet1 = pd.read_excel(excel_file1, skiprows=2)
    
    def Cell_Band_Name(raw):


        if (str(raw['antenna-model-number']).startswith("MX") or str(raw['antenna-model-number']).startswith("FF") or str(raw['antenna-model-number']).startswith("KE")) and str(raw['maximum-tilt']).startswith("140"):
            return (raw['NE Name'] + "_LB_" + str(raw['Sector-id']))

        elif (str(raw['antenna-model-number']).startswith("MX") or str(raw['antenna-model-number']).startswith("FF") or str(raw['antenna-model-number']).startswith("KE")) and str(raw['maximum-tilt']).startswith("120"):
            return (raw['NE Name'] + "_MB_" + str(raw['Sector-id']))
        
        elif (str(raw['antenna-model-number']).startswith("FVV")) and (str(raw['maximum-tilt']).startswith("140") or str(raw['maximum-tilt']).startswith("160") or str(raw['maximum-tilt']).startswith("180")):
            return (raw['NE Name'] + "_LB_" + str(raw['Sector-id']))

        elif (str(raw['antenna-model-number']).startswith("FVV")) and str(raw['maximum-tilt']).startswith("120"):
            return (raw['NE Name'] + "_MB_" + str(raw['Sector-id']))
        
        elif str(raw['antenna-model-number']).startswith("120") and str(raw['maximum-tilt']).startswith("120"):
            return (raw['NE Name'] + "_LB_" + str(raw['Sector-id']))
        
        elif str(raw['antenna-model-number']).startswith("120") and str(raw['maximum-tilt']).startswith("100"):
            return (raw['NE Name'] + "_MB_" + str(raw['Sector-id']))
        
        elif str(raw['antenna-model-number']).startswith("-"):
            return ("ERROR")
        
        elif pd.isna(raw['antenna-model-number']):
            return ("ERROR")
        
        else :
            return ("NEED TO CHECK")



    def Sector(raw):
        raw['sectorid'] = (raw['o-ran-radio-unit-info/o-ran-ru-id'] % 10)
        return (raw['sectorid'])

    sheet1['Sector-id'] = sheet1.apply(Sector, axis=1)

    sheet1['user-label_New'] = sheet1.apply(Cell_Band_Name, axis=1)




    table = pd.DataFrame(sheet1)

    excel_output_path = directory_path + '/RETINFO_' + time + '.xlsx'
    writer = pd.ExcelWriter(excel_output_path, engine='openpyxl')

    # Write the table to the Excel file
    table.to_excel(writer, sheet_name='RET_INFO', index=False)


    # Access the XlsxWriter workbook and worksheet objects
    workbook = writer.book
    worksheet = writer.sheets['RET_INFO']


    # Define a function to apply background color to rows based on the 'user-label' column
    def color_rows_based_on_column():
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            row_values = [cell.value for cell in row]
            if row_values[4] == 'ERROR':
                fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                for cell in row:
                    cell.fill = fill

            elif row_values[4] == 'NEED TO CHECK':
                fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                for cell in row:
                    cell.fill = fill

    # Apply the row coloring function based on the 'user-label' column
    color_rows_based_on_column()



    # Assuming df is your DataFrame and 'column_name' is the name of the column
    unique_values = sheet1['NE Name'].unique()

    # Create a list to store the repeated values
    repeated_values = []

    concatenated_values = []

    # Repeat each unique value three times and store them along with series numbers
    for value in unique_values:
        repeated_values.extend([(value, 1), (value, 2), (value, 3)])

    # Create a new DataFrame from the repeated values
    result_df = (pd.DataFrame(repeated_values, columns=['NE Name', 'Series_Number'])).sort_values(by='NE Name')

    for value in unique_values:
        for series_number in range(1, 4):
            concatenated_values.append(f"{value}_LB_{series_number}")

    result_df['user-label_New'] = pd.DataFrame(concatenated_values, columns=['Concatenated_Value'])

    concatenated_values = []
    for value in unique_values:
        for series_number in range(1, 4):
            concatenated_values.append(f"{value}_MB_{series_number}")


    result_df['MB'] = pd.DataFrame(concatenated_values, columns=['Concatenated_Value'])



    merged_df = pd.merge(result_df, table[['user-label_New', 'current-tilt']], on='user-label_New', how='left')

    merged_df.rename(columns={'current-tilt': 'LB_current-tilt'}, inplace=True)

    merged_df.drop(columns=['user-label_New'], inplace=True)


    merged_df.rename(columns={'MB': 'user-label_New'}, inplace=True)


    merged_df = pd.merge(merged_df, table[['user-label_New', 'current-tilt']], on='user-label_New', how='left')

    merged_df.rename(columns={'current-tilt': 'MB_current-tilt'}, inplace=True)

    merged_df.drop(columns=['user-label_New'], inplace=True)


    merged_df = merged_df.drop_duplicates()


    def TiltCompare(raw):
        if (raw['MB_current-tilt'] > raw['LB_current-tilt']):
            return ("NEED TO CHECK")
        elif (raw['MB_current-tilt'] <= raw['LB_current-tilt']):
            return ("GOOD")
        else :
            return ("RET MISSSING")
        

    merged_df['MB vs LB'] = merged_df.apply(TiltCompare, axis=1)

    def KEY(raw):
            return (raw['NE Name'] + "_" + str(raw['Series_Number']))
        
    merged_df['KEY'] = merged_df.apply(KEY, axis=1)

    merged_df = (merged_df.sort_values(by='KEY'))

    merged_df.to_excel(writer, sheet_name='MB vs LB TILT', index=False)

    workbook = writer.book
    worksheet = writer.sheets['MB vs LB TILT']


    # Define a function to apply background color to rows based on the 'user-label' column
    def RET_color_rows_based_on_column():
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            row_values = [cell.value for cell in row]
            if row_values[4] == 'RET MISSSING':
                fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                for cell in row:
                    cell.fill = fill

            elif row_values[4] == 'NEED TO CHECK':
                fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                for cell in row:
                    cell.fill = fill

    # Apply the row coloring function based on the 'user-label' column
    RET_color_rows_based_on_column()
    # Save the Excel file
    writer.close()


    messagebox.showinfo("Success", f"Updated file saved as:\n{excel_output_path}")


def browse_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xls")])
    if file_path:
        entry_file_path.delete(0, tk.END)
        entry_file_path.insert(0, file_path)

def run_processing():
    file_path = entry_file_path.get()
    if file_path:
        process_file(file_path)
    else:
        messagebox.showerror("Error", "Please select a file first.")

# Create UI window
root = tk.Tk()
root.title("LB/MB RET Compare")
root.geometry("500x200")
root.resizable(False, False)  # **Fix window size**

# tk.Label(root, text="Select an Excel file:").pack(pady=5)
# entry_file_path = tk.Entry(root, width=50)
# entry_file_path.pack()

# tk.Button(root, text="Browse", command=browse_file).pack(pady=5)
# tk.Button(root, text="Run", command=run_processing).pack(pady=10)


# File selection field
tk.Label(root, text="Select Excel File:", font=("Arial", 12, "bold")).grid(row=0, column=0, columnspan=2, padx=10, pady=5, sticky="w")
entry_file_path = tk.Entry(root, width=60)
entry_file_path.grid(row=1, column=0, columnspan=2, padx=15, pady=5, sticky="w")
tk.Button(root, text="Browse", command=browse_file,font=("Arial", 12, "bold")).grid(row=1, column=2, padx=10, pady=5)

# Run button
tk.Button(root, text="Run", command=run_processing, bg="green", fg="white", font=("Arial", 12, "bold")).grid(row=3, column=0, columnspan=3, pady=10)


# **Organizing Task & Vendor Selection**
frame = tk.Frame(root)
frame.grid(row=2, column=0, columnspan=2, padx=10, pady=5, sticky="w")

root.mainloop()