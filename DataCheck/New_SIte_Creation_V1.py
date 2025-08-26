import pandas as pd
from tkinter import Tk, filedialog, messagebox, StringVar
from tkinter import ttk
import openpyxl
from openpyxl.styles import Font, Alignment

# --- Helper Functions ---

def create_sample_reference_excel():
    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                             filetypes=[("Excel files", "*.xlsx")],
                                             title="Save Reference Excel File As")
    if not save_path: return
    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(title="Sheet1")
        headers = ["Site ID", "New Site ID", "Site Name", "New Latitude", "New Longitude", "RAD", "Description"]
        data_rows = [
            ["CLCLT00001A", "CLCLT30001A", 411981, 35.224473, -80.843835, 30, "Small Cell"],
            ["CLCLT00003A", "CLCLT30002A", 411983, 35.223557, -80.846035, 40, "Small Cell"],
            ["CLCLT00005A", "CLCLT00303A", 411984, 35.225466, -80.848106, 130, "Macro Cell"],
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = len(header) + 5
        for row_idx, row_data in enumerate(data_rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        wb.save(save_path)
        messagebox.showinfo("Success", f"Excel file created at:\n{save_path}")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to create file:\n{e}")

def check_required_columns(sheet_name, df, required_cols):
    missing = [col for col in required_cols if col not in df.columns]
    if missing:
        messagebox.showerror("Missing Columns", f"Sheet '{sheet_name}' is missing the following columns:\n\n" + "\n".join(missing))
        return False
    return True

# --- UI Interaction Functions ---

def select_file(path_variable, title):
    file_path = filedialog.askopenfilename(title=title, filetypes=[("Excel files", "*.xlsx",), ("All files", "*.*")])
    if file_path:
        path_variable.set(file_path)
    check_and_enable_start()

def check_and_enable_start():
    if ref_file_path.get() and planet_file_path.get():
        start_button.config(state="normal")
    else:
        start_button.config(state="disabled")

# --------- Main Processing Logic ---------
def start_processing():
    start_button.config(state="disabled")
    ref_browse_button.config(state="disabled")
    planet_browse_button.config(state="disabled")
    sample_button.config(state="disabled")
    
    reference_file = ref_file_path.get()
    multi_sheet_file = planet_file_path.get()

    try:
        # UPDATED: The progress bar now fills the horizontal space
        progress_bar.pack(fill='x', padx=10, pady=10)
        root.update_idletasks()

        xls = pd.ExcelFile(multi_sheet_file)
        sheets_to_process = {sheet_name: pd.read_excel(xls, sheet_name=sheet_name) for sheet_name in xls.sheet_names}
        
        total_steps = 2
        total_steps += sum(len(df) for name, df in sheets_to_process.items() if 'Site ID' in df.columns)
        total_steps += len(sheets_to_process)
        progress_bar['maximum'] = total_steps
        current_step = 0

        # UPDATED: Removed the "message" parameter as it's no longer needed
        def update_progress(steps=1):
            nonlocal current_step
            current_step += steps
            progress_bar['value'] = current_step
            root.update_idletasks()

        update_progress()
        reference_df = pd.read_excel(reference_file)
        site_id_mapping = reference_df.set_index('Site ID').to_dict(orient='index')
        
        for sheet_name, sheet_df in sheets_to_process.items():
            required_cols_map = {"Sites": ["Site ID"], "Sectors": ["Site ID"], "Groups": ["Name"]}
            if sheet_name in required_cols_map and not check_required_columns(sheet_name, sheet_df, required_cols_map[sheet_name]):
                return

            if 'Site ID' in sheet_df.columns:
                for idx, row in sheet_df.iterrows():
                    site_id = row['Site ID']
                    if site_id in site_id_mapping:
                        new_values = site_id_mapping[site_id]
                        sheet_df.at[idx, 'Site ID'] = new_values['New Site ID']
                        optional_cols = ['Latitude', 'Longitude', 'Description', 'Site Name', 'Site Name 2']
                        if all(col in sheet_df.columns for col in optional_cols):
                            sheet_df.at[idx, 'Latitude'], sheet_df.at[idx, 'Longitude'], sheet_df.at[idx, 'Description'], sheet_df.at[idx, 'Site Name'], sheet_df.at[idx, 'Site Name 2'] = new_values['New Latitude'], new_values['New Longitude'], new_values['Description'], new_values['Site Name'], f"{new_values['Site Name']}-{new_values['RAD']}"
                        if 'Height (ft)' in sheet_df.columns: sheet_df.at[idx, 'Height (ft)'] = new_values['RAD']
                        if 'Latitude' in sheet_df.columns and 'Longitude' in sheet_df.columns: sheet_df.at[idx, 'Latitude'], sheet_df.at[idx, 'Longitude'] = new_values['New Latitude'], new_values['New Longitude']
                    update_progress()

            if sheet_name == "Sites": sheet_df.loc[:, ["Custom: gNodeB_Id", "Custom: Cluster_ID", "Custom: gNodeB_Name", "Custom: gNodeB_Site_Number", "Custom: MLA_Partner", "Custom: TAC", "Custom: BEDC_CU_Number", "Custom: K8_ID_CUs", "Site UID"]] = ""
            if sheet_name == "Sectors": sheet_df.loc[:, ["Custom: Local_Cell_Id", "Custom: NR_Cell_Global_Id", "Custom: NR_Cell_Id", "Custom: NR_Cell_Name"]] = ""
            if sheet_name == "NR_Base_Stations": sheet_df.loc[:, "gNodeB ID"] = ""
            if sheet_name == "NR_Sector_Carriers": sheet_df.loc[:, ["Cell Name", "NCI", "TAC", "Cell ID", "Physical Cell ID", "First Zadoff Chu Sequence"]] = ["", "", 0, "", "", 0]
            if sheet_name == "NR_Sectors": sheet_df.loc[:, ["Physical Cell ID", "First Zadoff Chu Sequence"]] = ["", 0]
            
            update_progress()
        
        output_file = filedialog.asksaveasfilename(title="Save modified Excel file", defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not output_file: return

        with pd.ExcelWriter(output_file) as writer:
            for sheet_name, sheet_df in sheets_to_process.items():
                sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        update_progress()
        messagebox.showinfo("Success", f'Modified Excel saved as:\n{output_file}')

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred:\n{e}")
    
    finally:
        progress_bar.pack_forget()
        start_button.config(state="normal")
        ref_browse_button.config(state="normal")
        planet_browse_button.config(state="normal")
        sample_button.config(state="normal")
        check_and_enable_start()

# --------- Main GUI Layout ---------
root = Tk()
root.title("Site Creating Tools v2")
root.geometry("800x300")
root.minsize(700, 300)

ref_file_path = StringVar()
planet_file_path = StringVar()

top_frame = ttk.Frame(root)
top_frame.pack(side="top", fill="x", padx=10, pady=5)
sample_button = ttk.Button(top_frame, text="Download Sample File", command=create_sample_reference_excel)
sample_button.pack(side="right")

main_frame = ttk.Frame(root, padding="15")
main_frame.pack(fill="both", expand=True)

controls_frame = ttk.Frame(main_frame)
controls_frame.pack(fill="x", expand=False)

ref_frame = ttk.Frame(controls_frame)
ref_frame.pack(fill='x', expand=True, pady=5)
ttk.Label(ref_frame, text="Reference File:", width=15).pack(side="left")
ref_entry = ttk.Entry(ref_frame, textvariable=ref_file_path, state="readonly")
ref_entry.pack(side="left", fill="x", expand=True, padx=5)
ref_browse_button = ttk.Button(ref_frame, text="Browse...", command=lambda: select_file(ref_file_path, "Select the Reference Excel file"))
ref_browse_button.pack(side="left")

planet_frame = ttk.Frame(controls_frame)
planet_frame.pack(fill='x', expand=True, pady=(15, 5)) 
ttk.Label(planet_frame, text="PlanetExport File:", width=15).pack(side="left")
planet_entry = ttk.Entry(planet_frame, textvariable=planet_file_path, state="readonly")
planet_entry.pack(side="left", fill="x", expand=True, padx=5)
planet_browse_button = ttk.Button(planet_frame, text="Browse...", command=lambda: select_file(planet_file_path, "Select the PlanetExport Excel file"))
planet_browse_button.pack(side="left")

action_frame = ttk.Frame(controls_frame)
action_frame.pack(fill='x', expand=True, pady=(25, 0))
start_button = ttk.Button(action_frame, text="Start Processing", command=start_processing, state="disabled")
start_button.pack() 

# UPDATED: Shorter frame, as it only needs to hold the progress bar
progress_frame = ttk.Frame(main_frame, height=40)
progress_frame.pack(fill='x', expand=False, pady=(15, 0))
progress_frame.pack_propagate(False)

# UPDATED: Removed the progress_label creation
progress_bar = ttk.Progressbar(progress_frame, orient='horizontal', mode='determinate')

root.mainloop()