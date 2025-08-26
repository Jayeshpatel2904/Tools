import pandas as pd
from openpyxl import Workbook, load_workbook
from tkinter import filedialog
import os
import tkinter as tk
import datetime
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

time = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")

# Load data from the first Excel file (Sheet1)
excel_file1 = filedialog.askopenfilename()
# excel_file2 = filedialog.askopenfilename()

# columns_to_keep2 = ['timestamp','Custom: NR_Cell_Global_Id','RRC_ConnEstabAtt_Sum','rlc_vol_dl']
sheet1 = pd.read_excel(excel_file1, skiprows=2)
# sheet1 = pd.read_excel(excel_file1)

directory_path = os.path.dirname(excel_file1)

def Cell_Band_Name(raw):
    raw['sectorid'] = str(raw['o-ran-radio-unit-info/o-ran-ru-id'] % 10)

    if (str(raw['antenna-model-number']).startswith("MX") or str(raw['antenna-model-number']).startswith("FF") or str(raw['antenna-model-number']).startswith("KE")) and raw['maximum-tilt'] == "140":
        return (raw['NE Name'] + "_LB_" + raw['sectorid'])

    elif (str(raw['antenna-model-number']).startswith("MX") or str(raw['antenna-model-number']).startswith("FF") or str(raw['antenna-model-number']).startswith("KE")) and raw['maximum-tilt'] == "120":
        return (raw['NE Name'] + "_MB_" + raw['sectorid'])
    
    elif (str(raw['antenna-model-number']).startswith("FVV")) and (raw['maximum-tilt'] == "140" or raw['maximum-tilt'] == "160" or raw['maximum-tilt'] == "180"):
        return (raw['NE Name'] + "_LB_" + raw['sectorid'])

    elif (str(raw['antenna-model-number']).startswith("FVV")) and raw['maximum-tilt'] == "120":
        return (raw['NE Name'] + "_MB_" + raw['sectorid'])
    
    elif str(raw['antenna-model-number']).startswith("120") and raw['maximum-tilt'] == "120":
        return (raw['NE Name'] + "_LB_" + raw['sectorid'])
    
    elif str(raw['antenna-model-number']).startswith("120") and raw['maximum-tilt'] == "100":
        return (raw['NE Name'] + "_MB_" + raw['sectorid'])
    
    elif str(raw['antenna-model-number']).startswith("-"):
        return ("ERROR")
    
    elif pd.isna(raw['antenna-model-number']):
        return ("ERROR")
    
    else :
        return ("NEED TO CHECK")
    

def Band(raw):
    raw['sectorid'] = str(raw['o-ran-radio-unit-info/o-ran-ru-id'] % 10)

    if (str(raw['antenna-model-number']).startswith("MX") or str(raw['antenna-model-number']).startswith("FF") or str(raw['antenna-model-number']).startswith("KE")) and raw['maximum-tilt'] == "140":
        return ("LB")

    elif (str(raw['antenna-model-number']).startswith("MX") or str(raw['antenna-model-number']).startswith("FF") or str(raw['antenna-model-number']).startswith("KE")) and raw['maximum-tilt'] == "120":
        return ("MB")
    
    elif (str(raw['antenna-model-number']).startswith("FVV")) and (raw['maximum-tilt'] == "140" or raw['maximum-tilt'] == "160" or raw['maximum-tilt'] == "180"):
        return ("LB")

    elif (str(raw['antenna-model-number']).startswith("FVV")) and raw['maximum-tilt'] == "120":
        return ("MB")
    
    elif str(raw['antenna-model-number']).startswith("120") and raw['maximum-tilt'] == "120":
        return ("LB")
    
    elif str(raw['antenna-model-number']).startswith("120") and raw['maximum-tilt'] == "100":
        return ("MB")
    
    elif str(raw['antenna-model-number']).startswith("-"):
        return ("ERROR")
    
    elif pd.isna(raw['antenna-model-number']):
        return ("ERROR")
    
    else :
        return ("NEED TO CHECK")

def Sector(raw):
    raw['sectorid'] = str(raw['o-ran-radio-unit-info/o-ran-ru-id'] % 10)
    return (raw['sectorid'])


sheet1['user-label'] = sheet1.apply(Cell_Band_Name, axis=1)

sheet1['antenna-line-device-info/antenna-line-device-id'] = sheet1.apply(Band, axis=1)

sheet1['antenna-id'] = sheet1.apply(Sector, axis=1)

table = pd.DataFrame(sheet1)

excel_output_path = directory_path + '/RETINFO_' + time + '.xlsx'
writer = pd.ExcelWriter(excel_output_path, engine='openpyxl')

# Write the table to the Excel file
table.to_excel(writer, sheet_name='Sheet1', index=False)

# Access the XlsxWriter workbook and worksheet objects
workbook = writer.book
worksheet = writer.sheets['Sheet1']

# Define a function to apply background color to rows based on the 'user-label' column
def color_rows_based_on_column():
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        row_values = [cell.value for cell in row]
        if row_values[4] == 'ERROR':
            fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            for cell in row:
                cell.fill = fill

        elif row_values[4] == 'NEED TO CHECK':
            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            for cell in row:
                cell.fill = fill

# Apply the row coloring function based on the 'user-label' column
color_rows_based_on_column()

# Save the Excel file
writer.close()