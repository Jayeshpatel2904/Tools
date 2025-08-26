import pandas as pd
from tkinter import Tk, Button, filedialog, messagebox
import openpyxl
from openpyxl.styles import Font, Alignment

# --------- Create Sample Excel ---------
def create_sample_reference_excel():
    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                             filetypes=[("Excel files", "*.xlsx")],
                                             title="Save Reference Excel File As")

    if save_path:
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            ws = wb.create_sheet(title="Sheet1")

            headers = [
                "Site ID", "New Site ID", "Site Name",
                "New Latitude", "New Longitude", "RAD", "Description"
            ]

            data_rows = [
                ["CLCLT00001A", "CLCLT30001A", 411981, 35.224473, -80.843835, 30, "Small Cell"],
                ["CLCLT00003A", "CLCLT30002A", 411983, 35.223557, -80.846035, 40, "Small Cell"],
                ["CLCLT00005A", "CLCLT00303A", 411984, 35.225466, -80.848106, 130, "Macro Cell"],
            ]

            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = len(header) + 5

            for row_idx, row_data in enumerate(data_rows, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            wb.save(save_path)
            messagebox.showinfo("Success", f"Excel file created at:\n{save_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to create file:\n{e}")

# --------- Column Check ---------
def check_required_columns(sheet_name, df, required_cols):
    missing = [col for col in required_cols if col not in df.columns]
    if missing:
        messagebox.showerror("Missing Columns",
                             f"Sheet '{sheet_name}' is missing the following columns:\n\n" +
                             "\n".join(missing))
        return False
    return True

# --------- Main Processing ---------
def run_excel_processing():
    file_root = Tk()
    file_root.withdraw()
    reference_file = filedialog.askopenfilename(title="Select the reference Excel file",
                                                 filetypes=[("Excel files", "*.xlsx")])
    file_root.destroy()
    if not reference_file:
        messagebox.showerror("Error", "No reference file selected!")
        return

    reference_df = pd.read_excel(reference_file)
    site_id_mapping = reference_df.set_index('Site ID').to_dict(orient='index')

    file_root = Tk()
    file_root.withdraw()
    multi_sheet_file = filedialog.askopenfilename(title="Select the PlanetExport Excel file",
                                                  filetypes=[("Excel files", "*.xlsx")])
    file_root.destroy()
    if not multi_sheet_file:
        messagebox.showerror("Error", "No PlanetExport file selected!")
        return

    multi_sheet_excel = pd.read_excel(multi_sheet_file, sheet_name=None)

    for sheet_name, sheet_df in multi_sheet_excel.items():

        if sheet_name == "Sites":
            required_cols = ["Site ID"]
            if not check_required_columns(sheet_name, sheet_df, required_cols):
                return

        if sheet_name == "Sectors":
            required_cols = ["Site ID"]
            if not check_required_columns(sheet_name, sheet_df, required_cols):
                return

        if sheet_name == "Groups":
            required_cols = ["Name"]
            if not check_required_columns(sheet_name, sheet_df, required_cols):
                return

        if sheet_name == "NR_Sectors":
            required_cols = []
            if not check_required_columns(sheet_name, sheet_df, required_cols):
                return

        if sheet_name == "NR_Sector_Carriers":
            required_cols = []
            if not check_required_columns(sheet_name, sheet_df, required_cols):
                return

        if sheet_name == "NR_Base_Stations":
            required_cols = []
            if not check_required_columns(sheet_name, sheet_df, required_cols):
                return

        if 'Site ID' in sheet_df.columns:
            for idx, row in sheet_df.iterrows():
                site_id = row['Site ID']
                if site_id in site_id_mapping:
                    new_values = site_id_mapping[site_id]
                    sheet_df.at[idx, 'Site ID'] = new_values['New Site ID']

                    # Optional fields
                    optional_cols = ['Latitude', 'Longitude', 'Description', 'Site Name', 'Site Name 2']
                    if all(col in sheet_df.columns for col in optional_cols):
                        sheet_df.at[idx, 'Latitude'] = new_values['New Latitude']
                        sheet_df.at[idx, 'Longitude'] = new_values['New Longitude']
                        sheet_df.at[idx, 'Description'] = new_values['Description']
                        sheet_df.at[idx, 'Site Name'] = new_values['Site Name']
                        sheet_df.at[idx, 'Site Name 2'] = f"{new_values['Site Name']}-{new_values['RAD']}"

                    if 'Height (ft)' in sheet_df.columns:
                        sheet_df.at[idx, 'Height (ft)'] = new_values['RAD']

                    if 'Latitude' in sheet_df.columns and 'Longitude' in sheet_df.columns:
                        sheet_df.at[idx, 'Latitude'] = new_values['New Latitude']
                        sheet_df.at[idx, 'Longitude'] = new_values['New Longitude']


        if sheet_name == "Sites":
            for col in [
                "Custom: gNodeB_Id", "Custom: Cluster_ID", "Custom: gNodeB_Name",
                "Custom: gNodeB_Site_Number", "Custom: MLA_Partner", "Custom: TAC",
                "Custom: BEDC_CU_Number", "Custom: K8_ID_CUs", "Site UID"
            ]:
                sheet_df[col] = ""

        if sheet_name == "Sectors":
            for col in [
                "Custom: Local_Cell_Id", "Custom: NR_Cell_Global_Id", "Custom: NR_Cell_Id",
                "Custom: NR_Cell_Name"
            ]:
                sheet_df[col] = ""
            # sheet_df["Group: PLAN_3YEAR"] = True  # Uncomment if needed

        if sheet_name == "NR_Base_Stations":
            sheet_df["gNodeB ID"] = ""

        if sheet_name == "NR_Sector_Carriers":
            sheet_df["Cell Name"] = ""
            sheet_df["NCI"] = ""
            sheet_df["TAC"] = 0
            sheet_df["Cell ID"] = ""
            sheet_df["Physical Cell ID"] = ""
            sheet_df["First Zadoff Chu Sequence"] = 0

        if sheet_name == "NR_Sectors":
            sheet_df["Physical Cell ID"] = ""
            sheet_df["First Zadoff Chu Sequence"] = 0

    file_root = Tk()
    file_root.withdraw()
    output_file = filedialog.asksaveasfilename(title="Save modified Excel file",
                                               defaultextension=".xlsx",
                                               filetypes=[("Excel files", "*.xlsx")])
    file_root.destroy()
    if not output_file:
        messagebox.showerror("Error", "No output file selected!")
        return

    with pd.ExcelWriter(output_file) as writer:
        for sheet_name, sheet_df in multi_sheet_excel.items():
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)

    messagebox.showinfo("Success", f'Modified Excel saved as:\n{output_file}')

# --------- Main GUI ---------
root = Tk()
root.title("Site Creating Tools")
root.geometry("500x300")
root.resizable(False, False)

Button(root, text="Download Sample Reference Excel", command=create_sample_reference_excel,
       height=2, width=55).pack(pady=40)

Button(root, text="Create New Site", command=run_excel_processing,
       height=2, width=55).pack(pady=40)

root.mainloop()
