from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.chart import BarChart, Reference, LineChart
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference, LineChart
from tkinter import filedialog
import os
import tkinter as tk
import datetime
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows

time = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")

# Load data from the first Excel file (Sheet1)
excel_file1 = filedialog.askopenfilename()
excel_file2 = filedialog.askopenfilename()

# columns_to_keep2 = ['timestamp','Custom: NR_Cell_Global_Id','RRC_ConnEstabAtt_Sum','rlc_vol_dl']
sheet1 = pd.read_excel(excel_file1)

directory_path = os.path.dirname(excel_file1)



# if sheet1['timestamp'].str.len() == 10:
sheet1['timestamp'] = sheet1['timestamp'].astype(str).str[4:5] + '/' + sheet1['timestamp'].astype(str).str[:2] + '/' + sheet1['timestamp'].astype(str).str[6:]


sheet1.rename(columns={'label.NRCGI': 'Custom: NR_Cell_Global_Id'}, inplace=True)


# Load data from the second Excel file (specific tab and column)
# excel_file2 = filedialog.askopenfilename()
sheet2_tab = 'Sectors'  # Change to the name of the tab in file2.xlsx
column_to_join_from_sheet2 = 'Custom: NR_Cell_Global_Id'  # Change to the column name in sheet2.xlsx

columns_to_keep1 = ['Custom: NR_Cell_Global_Id', 'Custom: NR_Cell_Name']

# Read only the specified column from the second Excel file
sheet2_column = pd.read_excel(excel_file2, sheet_name=sheet2_tab, usecols=columns_to_keep1)

print(sheet2_column)

# columns_to_keep1 = ['Custom: NR_Cell_Global_Id', 'Custom: NR_Cell_Name']
columns_to_keep2 = ['timestamp', 'RRC_ConnEstabAtt_Sum','rlc_vol_dl']

# Merge the data

# merged_data = sheet1.merge(sheet2_column, left_on='Custom: NR_Cell_Global_Id', right_index=True)

result = pd.merge(sheet1, sheet2_column, on='Custom: NR_Cell_Global_Id', how='inner')

# Remove selected columns
columns_to_remove = ['Custom: NR_Cell_Global_Id', 'Voice Accessibility %']  # List of column names to remove
df = result.drop(columns=columns_to_remove)

# pivot_table = pd.DataFrame(df)


# unique_values = df['Custom: NR_Cell_Name'].unique()

# df = pd.DataFrame(df)


import pandas as pd
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Get unique values from the 'Category' column
unique_values = df['Custom: NR_Cell_Name'].unique()


# Create a new Excel workbook
excel_output_path = directory_path + '/filtered_data_1.xlsx'
wb = Workbook()

# Write the original data to the 'Original' sheet
ws_original = wb.active
for row in dataframe_to_rows(df, index=False, header=True):
    ws_original.append(row)

# Iterate over unique values and create charts on new sheets
for unique_value in unique_values:
    filtered_df = df[df['Custom: NR_Cell_Name'] == unique_value]
    
    # Create a new worksheet
    ws = wb.create_sheet(title=unique_value)
    
    # Write the filtered data to the worksheet
    for row in dataframe_to_rows(filtered_df, index=False, header=True):
        ws.append(row)
    
    # Create a line chart
    chart = LineChart()
    chart.title = f'Chart for {unique_value}'
    
    # Set data ranges for the chart series
    data = Reference(ws, min_col=2, min_row=1, max_row=len(filtered_df) + 1, max_col=4)
    categories = Reference(ws, min_col=1, min_row=1, max_row=len(filtered_df) + 1)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    
    # Add the chart to the worksheet
    ws.add_chart(chart, 'E2')

# Save the modified Excel file
wb.save(excel_output_path)

print("Filtered data and charts written to Excel tabs.")